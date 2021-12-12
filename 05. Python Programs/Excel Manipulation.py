# ---------------------------------------------------------------------------------------------------------------------
# Excel Manipulation using Python:
# ---------------------------------------------------------------------------------------------------------------------

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('Transaction.xlsx')       # load excel workbook
print(wb.sheetnames)                            # get sheet names -> ['Sheet1', 'Sheet2']

current_sheet = wb['Sheet1']                    # access specific sheet

cell1 = current_sheet['a1']                     # access a cell
print(cell1)                                    # <Cell 'Sheet1'.A1>
print(cell1.value)

cell2 = current_sheet.cell(1, 2)
print(cell2)                                    # <Cell 'Sheet1'.B1>
print(cell2.value)

print('Max Row: ')
print(current_sheet.max_row)                    # number of rows in the sheet

# Loop through all the rows
print('\n')
for row in range(2, current_sheet.max_row + 1):
    cell = current_sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = current_sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price


# ---------------------------------------------------------------------------------------------------------------------
# Add bar chart
# ---------------------------------------------------------------------------------------------------------------------

# select a range of values using Reference class
values = Reference(current_sheet,  # values is the object holding all the values in 4th column
                   min_row=2,
                   max_row=current_sheet.max_row,
                   min_col=3,
                   max_col=4)

bar_chart = BarChart()        # bar_chart is an object of BarChart class
bar_chart.add_data(values)
current_sheet.add_chart(bar_chart, 'b7')

wb.save('Transactions2.xlsx')
